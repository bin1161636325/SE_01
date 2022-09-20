# import os
import asyncio
import re
# import time
import openpyxl
# import random

from pyppeteer import launch
from bs4 import BeautifulSoup

# 港澳台
HongKong = []
Macao = []
Taiwan = []

# 请求网页
async def pyppteer_fetchUrl(url):
    browser = await launch({'headless': False,'dumpio':True, 'autoClose':True})
    page = await browser.newPage()
    await page.goto(url, timeout=10000000)
    # time.sleep(10)
    # 等待
    await asyncio.wait([page.waitForNavigation()])
    str = await page.content()
    await browser.close()
    return str

def fetchUrl(url):
    return asyncio.get_event_loop().run_until_complete(pyppteer_fetchUrl(url))

# 获取每页的url
def getPageUrl():
    for page in range(1, 42):
        if page == 1:
            yield 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml'
        else:
            url = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd_'+ str(page) + '.shtml'
            yield url

# 获取标题的title、link、date
def getTitleUrl(html):
    bsobj = BeautifulSoup(html,'html.parser') # 解析网页
    titleList = bsobj.find('div', attrs={"class":"list"}).ul.find_all("li")
    for item in titleList:
        link = "http://www.nhc.gov.cn" + item.a["href"]
        title = item.a["title"]
        date = item.span.text
        yield title, link, date

# 用正则表达式匹配数据
def getContent(html, result, date):
    bsobj = BeautifulSoup(html, 'html.parser')
    cnt = bsobj.find('div', attrs={"id": "xw_box"}).find_all("p")
    s = ""
    if cnt:
        # 获取文档中的p标签的所有的叠加
        for item in cnt:
            s = s + item.text
        # 初始化
        Anhui = '0'
        Beijing = '0'
        Chongqing = '0'
        Fujian = '0'
        Gansu = '0'
        Guangdong = '0'
        Guangxi = '0'
        Guizhou = '0'
        Hainan = '0'
        Hebei = '0'
        Henan = '0'
        Heilongjiang = '0'
        Hubei = '0'
        Hunan = '0'
        Jiangxi = '0'
        Jilin = '0'
        Jiangsu = '0'
        Liaoning = '0'
        Inner_Mongoria = '0'
        Ningxia = '0'
        Qinghai = '0'
        Shanxi = '0'
        Shandong = '0'
        Shaanxi = '0'
        Shanghai = '0'
        Sichuan = '0'
        Tianjing = '0'
        Tibet = '0'
        Xinjiang = '0'
        Bingtuan = '0'
        Yunnan = '0'
        Zhejiang = '0'

        # 本土新增确诊
        newly_increased = re.search('31个省（自治区、直辖市）和新疆生产建设兵团报告新增确诊病例(.*?)本土病例(\d+)例（(.*?)）', s)
        if (newly_increased != None):
            # print(newly_increased.group(2))
            # print(newly_increased.group(3))
            newly_increased_num = newly_increased.group(2)
            # 判断特殊的新增
            try:
                all_in = re.search('在(.*)', newly_increased.group(3))
            except:
                all_in = '0'
            # 各省的新增
            # 安徽
            try:
                Anhui = re.search('安徽(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('安徽', all_in.group(1))
                    if flag != None:
                        Anhui = newly_increased_num
                except:
                    Anhui = '0'

            # 北京
            try:
                Beijing = re.search('北京(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('北京', all_in.group(1))
                    if flag != None:
                        Beijing = newly_increased_num
                except:
                    Beijing = '0'
            # 重庆
            try:
                Chongqing = re.search('重庆(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('重庆', all_in.group(1))
                    if flag != None:
                        Chongqing = newly_increased_num
                except:
                    Chongqing = '0'
            # 福建
            try:
                Fujian = re.search('福建(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('福建', all_in.group(1))
                    if flag != None:
                        Fujian = newly_increased_num
                except:
                    Fujian = '0'
            # 甘肃
            try:
                Gansu = re.search('甘肃(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('甘肃', all_in.group(1))
                    if flag != None:
                        Gansu = newly_increased_num
                except:
                    Gansu = '0'
            # 广东
            try:
                Guangdong = re.search('广东(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('广东', all_in.group(1))
                    if flag != None:
                        Guangdong = newly_increased_num
                except:
                    Guangdong = '0'
            # 广西
            try:
                Guangxi = re.search('广西(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('广西', all_in.group(1))
                    if flag != None:
                        Guangxi = newly_increased_num
                except:
                    Guangxi = '0'
            # 贵州
            try:
                Guizhou = re.search('贵州(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('贵州', all_in.group(1))
                    if flag != None:
                        Guizhou = newly_increased_num
                except:
                    Guizhou = '0'
            # 海南
            try:
                Hainan = re.search('海南(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('海南', all_in.group(1))
                    if flag != None:
                        Hainan = newly_increased_num
                except:
                    Hainan = '0'
            # 河北
            try:
                Hebei = re.search('河北(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('河北', all_in.group(1))
                    if flag != None:
                        Hebei = newly_increased_num
                except:
                    Hebei = '0'
            # 河南
            try:
                Henan = re.search('河南(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('河南', all_in.group(1))
                    if flag != None:
                        Henan = newly_increased_num
                except:
                    Henan = '0'
            # 黑龙江
            try:
                Heilongjiang = re.search('黑龙江(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('黑龙江', all_in.group(1))
                    if flag != None:
                        Heilongjiang = newly_increased_num
                except:
                    Heilongjiang = '0'
            # 湖北
            try:
                Hubei = re.search('湖北(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('湖北', all_in.group(1))
                    if flag != None:
                        Hubei = newly_increased_num
                except:
                    Hubei = '0'
            # 湖南
            try:
                Hunan = re.search('湖南(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('湖南', all_in.group(1))
                    if flag != None:
                        Hunan = newly_increased_num
                except:
                    Hunan = '0'
            # 江西
            try:
                Jiangxi = re.search('江西(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('江西', all_in.group(1))
                    if flag != None:
                        Jiangxi = newly_increased_num
                except:
                    Jiangxi = '0'
            # 吉林
            try:
                Jilin = re.search('吉林(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('吉林', all_in.group(1))
                    if flag != None:
                        Jilin = newly_increased_num
                except:
                    Jilin = '0'
            # 江苏
            try:
                Jiangsu = re.search('江苏(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('江苏', all_in.group(1))
                    if flag != None:
                        Jiangsu = newly_increased_num
                except:
                    Jiangsu = '0'
            # 辽宁
            try:
                Liaoning = re.search('辽宁(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('辽宁', all_in.group(1))
                    if flag != None:
                        Liaoning = newly_increased_num
                except:
                    Liaoning = '0'
            # 内蒙古
            try:
                Inner_Mongoria = re.search('内蒙古(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('内蒙古', all_in.group(1))
                    if flag != None:
                        Inner_Mongoria = newly_increased_num
                except:
                    Inner_Mongoria = '0'
            # 宁夏
            try:
                Ningxia = re.search('宁夏(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('宁夏', all_in.group(1))
                    if flag != None:
                        Ningxia = newly_increased_num
                except:
                    Ningxia = '0'
            # 青海
            try:
                Qinghai = re.search('青海(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('青海', all_in.group(1))
                    if flag != None:
                        Qinghai = newly_increased_num
                except:
                    Qinghai = '0'
            # 山西
            try:
                Shanxi = re.search('山西(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('山西', all_in.group(1))
                    if flag != None:
                        Shanxi = newly_increased_num
                except:
                    Shanxi = '0'
            # 山东
            try:
                Shandong = re.search('山东(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('山东', all_in.group(1))
                    if flag != None:
                        Shandong = newly_increased_num
                except:
                    Shandong = '0'
            # 陕西
            try:
                Shaanxi = re.search('陕西(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('陕西', all_in.group(1))
                    if flag != None:
                        Shaanxi = newly_increased_num
                except:
                    Shaanxi = '0'
            # 上海
            try:
                Shanghai = re.search('上海(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('上海', all_in.group(1))
                    if flag != None:
                        Shanghai = newly_increased_num
                except:
                    Shanghai = '0'
            # 四川
            try:
                Sichuan = re.search('四川(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('四川', all_in.group(1))
                    if flag != None:
                        Sichuan = newly_increased_num
                except:
                    Sichuan = '0'
            # 天津
            try:
                Tianjing = re.search('天津(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('天津', all_in.group(1))
                    if flag != None:
                        Tianjing = newly_increased_num
                except:
                    Tianjing = '0'
            # 西藏
            try:
                Tibet = re.search('西藏(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('西藏', all_in.group(1))
                    if flag != None:
                        Tibet = newly_increased_num
                except:
                    Tibet = '0'
            # 新疆
            try:
                Xinjiang = re.search('新疆(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('新疆', all_in.group(1))
                    if flag != None:
                        Xinjiang = newly_increased_num
                except:
                    Xinjiang = '0'
            # 兵团
            try:
                Bingtuan = re.search('兵团(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('兵团', all_in.group(1))
                    if flag != None:
                        Bingtuan = newly_increased_num
                except:
                    Bingtuan = '0'
            # 云南
            try:
                Yunnan = re.search('云南(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('云南', all_in.group(1))
                    if flag != None:
                        Yunnan = newly_increased_num
                except:
                    Yunnan = '0'
            # 浙江
            try:
                Zhejiang = re.search('浙江(\d+)例', newly_increased.group(3)).group(1)
            except:
                try:
                    flag = re.search('浙江', all_in.group(1))
                    if flag != None:
                        Zhejiang = newly_increased_num
                except:
                    Zhejiang = '0'

        else:
            # 本土无新增
            newly_increased_num = '0'
        # 港澳台
        try:
            GAT = re.search('累计收到港澳台地区通报确诊病例(\d+)例(.*?)香港特别行政区(\d+)例(.*?)澳门特别行政区(\d+)例(.*?)台湾地区(\d+)例', s)
            # 香港
            HongKong.append(str(GAT.group(3)))
            Macao.append(str(GAT.group(5)))
            Taiwan.append(str(GAT.group(7)))

        except:
            pass

        # 插入数据
        result.append([date, '新增确诊', int(newly_increased_num), int(Anhui), int(Beijing), int(Chongqing), int(Fujian), int(Gansu), int(Guangdong), int(Guangxi), int(Guizhou), int(Hainan), int(Hebei), int(Henan), int(Heilongjiang), int(Hubei), int(Hunan), int(Jiangxi), int(Jilin), int(Jiangsu), int(Liaoning), int(Inner_Mongoria), int(Ningxia), int(Qinghai), int(Shanxi), int(Shandong), int(Shaanxi), int(Shanghai), int(Sichuan), int(Tianjing), int(Tibet), int(Xinjiang) + int(Bingtuan), int(Yunnan), int(Zhejiang)])

        # 初始化
        Anhui = '0'
        Beijing = '0'
        Chongqing = '0'
        Fujian = '0'
        Gansu = '0'
        Guangdong = '0'
        Guangxi = '0'
        Guizhou = '0'
        Hainan = '0'
        Hebei = '0'
        Henan = '0'
        Heilongjiang = '0'
        Hubei = '0'
        Hunan = '0'
        Jiangxi = '0'
        Jilin = '0'
        Jiangsu = '0'
        Liaoning = '0'
        Inner_Mongoria = '0'
        Ningxia = '0'
        Qinghai = '0'
        Shanxi = '0'
        Shandong = '0'
        Shaanxi = '0'
        Shanghai = '0'
        Sichuan = '0'
        Tianjing = '0'
        Tibet = '0'
        Xinjiang = '0'
        Bingtuan = '0'
        Yunnan = '0'
        Zhejiang = '0'

        # 本土新增无症状
        newly_increased_asympto = re.search('31个省（自治区、直辖市）和新疆生产建设兵团报告新增无症状感染者(.*?)本土(\d+)例（(.*?)）', s)
        if (newly_increased_asympto != None):
            # print(newly_increased.group(2))
            # print(newly_increased.group(3))
            newly_increased_asympto_num = newly_increased_asympto.group(2)
            try:
                all_in = re.search('在(.*)', newly_increased_asympto.group(3))
                print(all_in.group(1))
            except:
                all_in = '0'
            # 各省的新增
            # 安徽
            try:
                Anhui = re.search('安徽(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('安徽', all_in.group(1))
                    if flag != None:
                        Anhui = newly_increased_asympto_num
                except:
                    Anhui = '0'

            # 北京
            try:
                Beijing = re.search('北京(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('北京', all_in.group(1))
                    if flag != None:
                        Beijing = newly_increased_asympto_num
                    print(Beijing)
                except:
                    Beijing = '0'
            # 重庆
            try:
                Chongqing = re.search('重庆(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('重庆', all_in.group(1))
                    if flag != None:
                        Chongqing = newly_increased_asympto_num
                except:
                    Chongqing = '0'
            # 福建
            try:
                Fujian = re.search('福建(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('福建', all_in.group(1))
                    if flag != None:
                        Fujian = newly_increased_asympto_num
                except:
                    Fujian = '0'
            # 甘肃
            try:
                Gansu = re.search('甘肃(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('甘肃', all_in.group(1))
                    if flag != None:
                        Gansu = newly_increased_asympto_num
                except:
                    Gansu = '0'
            # 广东
            try:
                Guangdong = re.search('广东(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('广东', all_in.group(1))
                    if flag != None:
                        Guangdong = newly_increased_asympto_num
                except:
                    Guangdong = '0'
            # 广西
            try:
                Guangxi = re.search('广西(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('广西', all_in.group(1))
                    if flag != None:
                        Guangxi = newly_increased_asympto_num
                except:
                    Guangxi = '0'
            # 贵州
            try:
                Guizhou = re.search('贵州(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('贵州', all_in.group(1))
                    if flag != None:
                        Guizhou = newly_increased_asympto_num
                except:
                    Guizhou = '0'
            # 海南
            try:
                Hainan = re.search('海南(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('海南', all_in.group(1))
                    if flag != None:
                        Hainan = newly_increased_asympto_num
                except:
                    Hainan = '0'
            # 河北
            try:
                Hebei = re.search('河北(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('河北', all_in.group(1))
                    if flag != None:
                        Hebei = newly_increased_asympto_num
                except:
                    Hebei = '0'
            # 河南
            try:
                Henan = re.search('河南(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('河南', all_in.group(1))
                    if flag != None:
                        Henan = newly_increased_asympto_num
                except:
                    Henan = '0'
            # 黑龙江
            try:
                Heilongjiang = re.search('黑龙江(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('黑龙江', all_in.group(1))
                    if flag != None:
                        Heilongjiang = newly_increased_asympto_num
                except:
                    Heilongjiang = '0'
            # 湖北
            try:
                Hubei = re.search('湖北(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('湖北', all_in.group(1))
                    if flag != None:
                        Hubei = newly_increased_asympto_num
                except:
                    Hubei = '0'
            # 湖南
            try:
                Hunan = re.search('湖南(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('湖南', all_in.group(1))
                    if flag != None:
                        Hunan = newly_increased_asympto_num
                except:
                    Hunan = '0'
            # 江西
            try:
                Jiangxi = re.search('江西(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('江西', all_in.group(1))
                    if flag != None:
                        Jiangxi = newly_increased_asympto_num
                except:
                    Jiangxi = '0'
            # 吉林
            try:
                Jilin = re.search('吉林(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('吉林', all_in.group(1))
                    if flag != None:
                        Jilin = newly_increased_asympto_num
                except:
                    Jilin = '0'
            # 江苏
            try:
                Jiangsu = re.search('江苏(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('江苏', all_in.group(1))
                    if flag != None:
                        Jiangsu = newly_increased_asympto_num
                except:
                    Jiangsu = '0'
            # 辽宁
            try:
                Liaoning = re.search('辽宁(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('辽宁', all_in.group(1))
                    if flag != None:
                        Liaoning = newly_increased_asympto_num
                except:
                    Liaoning = '0'
            # 内蒙古
            try:
                Inner_Mongoria = re.search('内蒙古(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('内蒙古', all_in.group(1))
                    if flag != None:
                        Inner_Mongoria = newly_increased_asympto_num
                except:
                    Inner_Mongoria = '0'
            # 宁夏
            try:
                Ningxia = re.search('宁夏(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('宁夏', all_in.group(1))
                    if flag != None:
                        Ningxia = newly_increased_asympto_num
                except:
                    Ningxia = '0'
            # 青海
            try:
                Qinghai = re.search('青海(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('青海', all_in.group(1))
                    if flag != None:
                        Qinghai = newly_increased_asympto_num
                except:
                    Qinghai = '0'
            # 山西
            try:
                Shanxi = re.search('山西(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('山西', all_in.group(1))
                    if flag != None:
                        Shanxi = newly_increased_asympto_num
                except:
                    Shanxi = '0'
            # 山东
            try:
                Shandong = re.search('山东(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('山东', all_in.group(1))
                    if flag != None:
                        Shandong = newly_increased_asympto_num
                except:
                    Shandong = '0'
            # 陕西
            try:
                Shaanxi = re.search('陕西(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('陕西', all_in.group(1))
                    if flag != None:
                        Shaanxi = newly_increased_asympto_num
                except:
                    Shaanxi = '0'
            # 上海
            try:
                Shanghai = re.search('上海(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('上海', all_in.group(1))
                    if flag != None:
                        Shanghai = newly_increased_asympto_num
                except:
                    Shanghai = '0'
            # 四川
            try:
                Sichuan = re.search('四川(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('四川', all_in.group(1))
                    if flag != None:
                        Sichuan = newly_increased_asympto_num
                except:
                    Sichuan = '0'
            # 天津
            try:
                Tianjing = re.search('天津(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('天津', all_in.group(1))
                    if flag != None:
                        Tianjing = newly_increased_asympto_num
                except:
                    Tianjing = '0'
            # 西藏
            try:
                Tibet = re.search('西藏(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('西藏', all_in.group(1))
                    if flag != None:
                        Tibet = newly_increased_asympto_num
                except:
                    Tibet = '0'
            # 新疆
            try:
                Xinjiang = re.search('新疆(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('新疆', all_in.group(1))
                    if flag != None:
                        Xinjiang = newly_increased_asympto_num
                except:
                    Xinjiang = '0'
            # 兵团
            try:
                Bingtuan = re.search('兵团(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('兵团', all_in.group(1))
                    if flag != None:
                        Bingtuan = newly_increased_asympto_num
                except:
                    Bingtuan = '0'
            # 云南
            try:
                Yunnan = re.search('云南(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('云南', all_in.group(1))
                    if flag != None:
                        Yunnan = newly_increased_asympto_num
                except:
                    Yunnan = '0'
            # 浙江
            try:
                Zhejiang = re.search('浙江(\d+)例', newly_increased_asympto.group(3)).group(1)
            except:
                try:
                    flag = re.search('浙江', all_in.group(1))
                    if flag != None:
                        Zhejiang = newly_increased_asympto_num
                except:
                    Zhejiang = '0'

        else:
            # 本土无新增
            newly_increased_asympto_num = '0'

        HongKong.append(0)
        Macao.append(0)
        Taiwan.append(0)

        result.append([date, '新增无症状', int(newly_increased_asympto_num), int(Anhui), int(Beijing), int(Chongqing), int(Fujian), int(Gansu), int(Guangdong), int(Guangxi), int(Guizhou), int(Hainan), int(Hebei), int(Henan), int(Heilongjiang), int(Hubei), int(Hunan), int(Jiangxi), int(Jilin), int(Jiangsu), int(Liaoning), int(Inner_Mongoria), int(Ningxia), int(Qinghai), int(Shanxi), int(Shandong), int(Shaanxi), int(Shanghai), int(Sichuan), int(Tianjing), int(Tibet), int(Xinjiang) + int(Bingtuan), int(Yunnan), int(Zhejiang)])

        return result

    return "爬取失败！"

# def saveFile(path, filename, content):
#     if not os.path.exists(path):
#         os.makedirs(path)
#
#     # 保存文件
#
#     with open(path + filename + ".txt", 'w', encoding='utf-8') as f:
#         f.write(content)

# 保存在excel
def save_to_excel(result):
    wb = openpyxl.Workbook()  # 建立一个Excel工作簿
    ws = wb.active  # 得到一个 sheet 的页面

    ws['A1'] = '时间'  # 直接给单元格赋值
    ws['B1'] = '新增确诊或无症状'
    ws['C1'] = '新增总人数'
    ws['D1'] = '安徽'
    ws['E1'] = '北京'
    ws['F1'] = '重庆'
    ws['G1'] = '福建'
    ws['H1'] = '甘肃'
    ws['I1'] = '广东'
    ws['J1'] = '广西'
    ws['K1'] = '贵州'
    ws['L1'] = '海南'
    ws['M1'] = '河北'
    ws['N1'] = '河南'
    ws['O1'] = '黑龙江'
    ws['P1'] = '湖北'
    ws['Q1'] = '湖南'
    ws['R1'] = '江西'
    ws['S1'] = '吉林'
    ws['T1'] = '江苏'
    ws['U1'] = '辽宁'
    ws['V1'] = '内蒙古'
    ws['W1'] = '宁夏'
    ws['X1'] = '青海'
    ws['Y1'] = '山西'
    ws['Z1'] = '山东'
    ws['AA1'] = '陕西'
    ws['AB1'] = '上海'
    ws['AC1'] = '四川'
    ws['AD1'] = '天津'
    ws['AE1'] = '西藏'
    ws['AF1'] = '新疆'
    ws['AG1'] = '云南'
    ws['AH1'] = '浙江'
    ws['AI1'] = '香港'
    ws['AJ1'] = '澳门'
    ws['AK1'] = '台湾'

    length = len(HongKong)
    for i in range(0, length - 2):
        if (i % 2 == 0):
            HongKong[i] = int(HongKong[i]) - int(HongKong[i + 2])
            Macao[i] = int(Macao[i]) - int(Macao[i + 2])
            Taiwan[i] = int(Taiwan[i]) - int(Taiwan[i + 2])

    for each in result:
        ws.append(each)  # 用此函数只能按行写入，从空白行开始

    for i in range(1, length + 1):
        ws.cell(row=i + 1, column=35).value = HongKong[i - 1]
        ws.cell(row=i + 1, column=36).value = Macao[i - 1]
        ws.cell(row=i + 1, column=37).value = Taiwan[i - 1]

    wb.save('疫情新增以及新增无症状.xlsx')

if "__main__" == __name__:
    result = []
    try:
        for url in getPageUrl():
            try:
                s = fetchUrl(url)
                for title,link,date in getTitleUrl(s):
                    print(title,link)
                    # 如果日期在1月21日之前，则直接退出
                    mon = int(date.split("-")[1])
                    day = int(date.split("-")[2])
                    html = fetchUrl(link)
                    # content = getContent(html, result)
                    title_date = re.search('截至(.*?)24时新型冠状病毒肺炎疫情最新情况', title).group(1)
                    if title_date == None:
                        continue
                    result = getContent(html, result, date.split("-")[0] + "年" + title_date)
                    # print(content)
                    # saveFile("../infor/", title, content)
                    print("-----"*20)
            except:
                try:
                    s = fetchUrl(url)
                    for title, link, date in getTitleUrl(s):
                        print(title, link)
                        # 如果日期在1月21日之前，则直接退出
                        mon = int(date.split("-")[1])
                        day = int(date.split("-")[2])
                        html = fetchUrl(link)
                        # content = getContent(html, result)
                        title_date = re.search('截至(.*?)24时新型冠状病毒肺炎疫情最新情况', title).group(1)
                        if title_date == None:
                            continue
                        result = getContent(html, result, date.split("-")[0] + "年" + title_date)
                        # print(content)
                        # saveFile("../infor/", title, content)
                        print("-----" * 20)
                except:
                    pass

        save_to_excel(result)
    except:
        save_to_excel(result)